import pandas as pd
import pyodbc
from openpyxl import load_workbook

"""
conn_str = (
    r'DRIVER={ODBC Driver 17 for SQL Server};'
    r'SERVER=10.10.2.25;'
    r'DATABASE=PCR;'
    r'Trusted_Connection=yes;'
)
conn = pyodbc.connect(conn_str)
cursor = conn.cursor() """

conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=10.10.2.25;DATABASE=PCR;UID=Pcrdwh;PWD=dPcrdwhV646!$W')
cursor = conn.cursor()

cursor.execute('''Select 
                        f.TipoCompra,
                        f.Meses, 
                        p.Estatus, 
                        p.Company as cliente,
                        f.MODELO, 
                        f.Kilometraje, 
                        f.CostoVehiculo, 
                        f.TotalCompra, 
                        f.IngresosAlquiler, 
                        f.OtrosIngresos, 
                        f.DiasRental, 
                        CASE 
                            WHEN f.DiasRental = 0 THEN NULL
                            ELSE (f.TotaldeIngresos) / (f.DiasRental)
                        END as revenue_per_day,
                        CASE 
                            WHEN f.Meses = 0 THEN NULL
                            ELSE (f.TotaldeIngresos) / (f.Meses)
                        END as revenue_per_month,
                        f.TOTALCOSTOS, 
                        f.INTERESES, 
                        f.SEGURO, 
                        f.DEPRECIACIONACUMULADA,
                        f.GastoAdmon
                    from 
                        VwFlotaAutosAF F LEFT OUTER JOIN PowerBI_Fleet P ON F.Unidad = P.UnitNumber 
                                        LEFT OUTER JOIN Tb_EstatusFlotaPCR E ON P.Estatus = E.Estatus''')

columns = [column[0] for column in cursor.description]
rows = cursor.fetchall()
df = pd.DataFrame.from_records(rows, columns=columns)

archivo_excel = 'simu.xlsx'
libro = load_workbook(archivo_excel)
hoja = libro['BD Ficha']  

"""max_row = hoja.max_row
for i in range(2, max_row + 1):  
    hoja.delete_rows(2)
print("Datos borrados " + int.max_row)"""

columnas_excel = ['F', 'H', 'M', 'N', 'P', 'S', 'T', 'V', 'W', 'X', 'Z', 'AA', 'AB', 'AH', 'AI', 'AJ', 'AK', 'BA']
fila_inicio = 2  

for index, row in df.iterrows():
    for col, val in zip(columnas_excel, row):
        hoja[f'{col}{fila_inicio + index}'].value = val

libro.save(archivo_excel)
libro.close()

cursor.close()
conn.close()

print("Los datos han sido escritos exitosamente en el archivo Excel.")
